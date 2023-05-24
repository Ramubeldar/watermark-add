# program for setting watermark

# importing important python libraries
import moviepy.editor as mp
from openpyxl import load_workbook
from moviepy.editor import *
import shutil
import os


# defining the function
def read_excel_row(row_num, excel_file):
    global font_first, font_second
    wb = load_workbook(filename=excel_file, read_only=True)
    worksheet = wb.active

    # row data preprocessing
    Font_Code = []
    for col in range(6, 7):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Font_Code.append(cell_value)
    # Split the single string element into a list of separate elements
    Font_Code = Font_Code[0].split('\n')

    # Create a new list to hold the elements without newline character
    Font_Codes = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Font_Code:
        Font_Codes.append(element.strip())

    Parent_folder_name = []
    for col in range(1, 2):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Parent_folder_name.append(cell_value)

    Input_File_Name = []
    for col in range(2, 3):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Input_File_Name.append(cell_value)

    Message = []
    for col in range(3, 4):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Message.append(cell_value)

    Duration = []
    for col in range(4, 5):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Duration.append(cell_value)

    Output_File_Name = []
    for col in range(5, 6):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Output_File_Name.append(cell_value)

    Parent_folder_name = Parent_folder_name[0].split('\n')

    # Create a new list to hold the elements without newline character
    Parent_folder_names = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Parent_folder_name:
        Parent_folder_names.append(element.strip())

    Input_File_Name = Input_File_Name[0].split('\n')

    # Create a new list to hold the elements without newline character
    Input_File_Names = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Input_File_Name:
        Input_File_Names.append(element.strip())

    # Split the single string element into a list of separate elements
    Message = Message[0].split('\n')

    # Create a new list to hold the elements without newline character
    Messages = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Message:
        Messages.append(element.strip())

    # Split the single string element into a list of separate elements
    Duration = Duration[0].split('\n')

    # Create a new list to hold the elements without newline character
    Durations = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Duration:
        Durations.append(element.strip())

    Output_File_Name = Output_File_Name[0].split('\n')

    # Create a new list to hold the elements without newline character
    Output_File_Names = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Output_File_Name:
        Output_File_Names.append(element.strip())
    try:
        video = mp.VideoFileClip(Input_File_Names[0])
    except:
        print("video file is not available for the row number: - ",row_num,Parent_folder_names[0])
        with open('error_log.txt', 'a') as f:
            data = str(Input_File_Names[0]) + "--->video file is not available for the row_number--->" + str(row_num)
            f.write(data + "\n")
        return
    clips = []

    # Create the watermark clip with the watermark text and set its position to top-left corner
    watermark_duration = video.duration
    watermark = (TextClip('EdutainmentVentures.com', fontsize=30, font='Comic-Sans-MS-Bold', color='white')
                 .set_position(('left', 'top'))
                 .set_duration(video.duration))
    watermark = watermark.set_opacity(0.35)
    # Animate the watermark by changing its position from top-left to left-right corner over the duration of the video
    watermark = watermark.set_position(
        lambda t: ('left', 'top') if t < watermark_duration / 2 else ('right', 'bottom'))

    def animate_position(t):
        speed = 20  # speed of the watermark
        pos = speed * t  # position of the watermark

        # Check if the watermark has hit the left or right edge of the video
        if pos < 0 or pos > 2 * (video.w - watermark.w):
            pos = abs(pos) % (2 * (video.w - watermark.w))  # set position to new value

        if pos > video.w - watermark.w:
            pos = 2 * (video.w - watermark.w) - pos  # reverse direction

        return pos, 0

    # Set the position of the watermark using the animate_position function
    watermark = watermark.set_position(animate_position)
    # Add the watermark and subtitle to the original video
    # create a CompositeVideoClip with the TextClips and the original video clip
    composite_clip = mp.CompositeVideoClip([video, watermark] + clips)

    # specify the name of the top-level folder
    top_level_folder_name = Parent_folder_names[0]

    # specify the name of the subfolder
    subfolder_name = Font_Codes[0]

    # check if the top-level folder already exists
    if not os.path.exists(top_level_folder_name):
        # create the top-level folder
        os.makedirs(top_level_folder_name)

    # check if the subfolder already exists
    subfolder_path = os.path.join(top_level_folder_name, subfolder_name)
    if os.path.exists(subfolder_path):
        # remove the subfolder and all its contents
        shutil.rmtree(subfolder_path)

    # create the subfolder within the top-level folder
    os.makedirs(subfolder_path)

    # Save the final edited video file in the child folder
    final_video_path = os.path.join(subfolder_path, Output_File_Names[0])

    composite_clip.write_videofile(final_video_path)
    print("video completed till the row number : - ",row_num)


# reading the exel file
excel_file_sheets = "Recipe1.xlsx"
workbook = load_workbook("Recipe1.xlsx")

# Select the active worksheet
worksheet = workbook.active

# Find the number of rows with data in excel sheet
num_rows = 0
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    if any(cell.value for cell in row):
        num_rows += 1

for row_number in range(15, 15+1):
    read_excel_row(row_number, excel_file_sheets)
